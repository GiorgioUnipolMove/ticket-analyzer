import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


# File con analisi manuali di riferimento (per confronto)
MANUAL_ANALYSIS_FILE = "Copia di Restituzione_device_aggiuntivi_ANALISI OPS.xlsx"
MANUAL_ANALYSIS_COL = "Analisi "  # nome colonna nel file manuale (con spazio)


class ExcelWriter:
    """Genera il file Excel di output con la colonna Analisi formattata."""

    HEADER_FILL = PatternFill('solid', fgColor='4472C4')
    HEADER_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    ANALISI_FILL = PatternFill('solid', fgColor='FFF2CC')
    MANUAL_FILL = PatternFill('solid', fgColor='D9E2F3')
    # Arancione per flag anomalie multi-ticket
    FLAG_FILL = PatternFill('solid', fgColor='FCE4D6')
    BODY_FONT = Font(name='Arial', size=10)

    def write(self, df_main: pd.DataFrame, results: dict, output_file: str,
              multi_ticket_flags: dict = None):
        """Genera il file Excel di output.

        Args:
            df_main: DataFrame principale (tab Restituzione_Device_Agg).
            results: Mappa ticket_id → testo analisi.
            output_file: Percorso del file di output.
            multi_ticket_flags: Mappa ticket_id → flag anomalia multi-ticket
                (opzionale, prodotto da RuleBasedAnalyzer).
        """
        if multi_ticket_flags is None:
            multi_ticket_flags = {}

        print(f"\nGenerazione output: {output_file}")

        df_out = df_main.copy()

        # Inserisci colonna Analisi dopo data_recesso_dispositivo
        try:
            col_idx = df_out.columns.get_loc('data_recesso_dispositivo') + 1
        except KeyError:
            col_idx = len(df_out.columns)
        df_out.insert(col_idx, 'Analisi', '')

        # Popola Analisi
        for idx, row in df_out.iterrows():
            tid = str(row.get('id_interaction', '')).strip()
            if tid in results:
                df_out.at[idx, 'Analisi'] = results[tid]

        # Aggiungi colonna Analisi Manuale (da file di confronto)
        manual_map = self._load_manual_analyses()
        analisi_col_idx = df_out.columns.get_loc('Analisi') + 1
        df_out.insert(analisi_col_idx, 'Analisi Manuale', '')
        if manual_map:
            for idx, row in df_out.iterrows():
                tid = str(row.get('id_interaction', '')).strip()
                if tid in manual_map:
                    df_out.at[idx, 'Analisi Manuale'] = manual_map[tid]
            manual_count = sum(1 for v in manual_map.values() if v)
            print(f"  📋 {manual_count} analisi manuali caricate per confronto")

        # Aggiungi colonna Flag Contratto (ticket multipli con esiti discordanti)
        manual_col_idx = df_out.columns.get_loc('Analisi Manuale') + 1
        df_out.insert(manual_col_idx, 'Flag Contratto', '')
        if multi_ticket_flags:
            flagged_count = 0
            for idx, row in df_out.iterrows():
                tid = str(row.get('id_interaction', '')).strip()
                if tid in multi_ticket_flags:
                    df_out.at[idx, 'Flag Contratto'] = multi_ticket_flags[tid]
                    flagged_count += 1
            print(f"  Ticket con flag contratto multiplo: {flagged_count}")

        # Rimuovi ClienteID dall'output
        if 'ClienteID' in df_out.columns:
            df_out = df_out.drop(columns=['ClienteID'])

        # Scrivi Excel
        df_out.to_excel(output_file, index=False, sheet_name='Analisi')

        # Formattazione
        self._format(output_file)

        filled = sum(1 for v in results.values() if v and not v.startswith('ERRORE'))
        print(f"  ✅ {filled} analisi scritte su {len(df_out)} righe")

    def _load_manual_analyses(self) -> dict:
        """Carica le analisi manuali dal file di confronto."""
        import os
        if not os.path.exists(MANUAL_ANALYSIS_FILE):
            print(f"  ⚠️  File confronto non trovato: {MANUAL_ANALYSIS_FILE}")
            return {}
        try:
            df = pd.read_excel(MANUAL_ANALYSIS_FILE, sheet_name=0, dtype=str)
            manual_map = {}
            if MANUAL_ANALYSIS_COL in df.columns:
                for _, row in df.iterrows():
                    tid = str(row.get('id_interaction', '')).strip()
                    val = str(row.get(MANUAL_ANALYSIS_COL, '')).strip()
                    if tid and val and val not in ('nan', '', 'NULL', 'None'):
                        manual_map[tid] = val
            return manual_map
        except Exception as e:
            print(f"  ⚠️  Errore lettura file confronto: {e}")
            return {}

    def _format(self, output_file: str):
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active

        # Headers
        for cell in ws[1]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        # Trova colonne Analisi, Analisi Manuale e Flag Contratto
        analisi_col = None
        manual_col = None
        flag_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'Analisi':
                analisi_col = idx
            elif cell.value == 'Analisi Manuale':
                manual_col = idx
            elif cell.value == 'Flag Contratto':
                flag_col = idx

        # Evidenzia colonna Analisi (giallo)
        if analisi_col:
            for row in ws.iter_rows(min_row=2, min_col=analisi_col, max_col=analisi_col):
                for cell in row:
                    cell.fill = self.ANALISI_FILL
                    cell.font = self.BODY_FONT

        # Evidenzia colonna Analisi Manuale (azzurro)
        if manual_col:
            for row in ws.iter_rows(min_row=2, min_col=manual_col, max_col=manual_col):
                for cell in row:
                    cell.fill = self.MANUAL_FILL
                    cell.font = self.BODY_FONT

        # Evidenzia colonna Flag Contratto (arancione) — solo celle non vuote
        if flag_col:
            for row in ws.iter_rows(min_row=2, min_col=flag_col, max_col=flag_col):
                for cell in row:
                    if cell.value:
                        cell.fill = self.FLAG_FILL
                    cell.font = self.BODY_FONT

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        wb.save(output_file)
